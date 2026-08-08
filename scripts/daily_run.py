#!/usr/bin/env python3
"""Run the daily XStudioz growth engine.

Reads source snapshots from ``data/raw/``, runs the full pipeline, writes the
brief to ``reports/`` and exits non-zero if the self-check gate blocked the
brief — so a scheduler can tell "ran and produced advice" from "ran and did
not trust its own output".

Usage
-----
    python3 scripts/daily_run.py                       # today, using latest snapshots
    python3 scripts/daily_run.py --date 2026-06-03     # a specific day
    python3 scripts/daily_run.py --dry-run             # do not write anything

Snapshots
---------
The fetching step is deliberately outside this script: a Claude Routine (or
any other job) drops text exports into ``data/raw/`` and then calls this. That
separation is what makes every run replayable offline.

    data/raw/orders/YYYY-MM-DD.md        order + daily-ledger workbook
    data/raw/inquiries/YYYY-MM-DD.md     inquiry workbook
    data/raw/gig/YYYY-MM-DD.json         scraped public gig signals
    data/raw/order_tracker/*.xlsx        CSR handoff tracker
"""

from __future__ import annotations

import argparse
import datetime as _dt
import json
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from xstudioz import ingest, pipeline, snapshot  # noqa: E402


def latest_snapshot(directory: Path, on_or_before: _dt.date,
                    suffixes: tuple[str, ...] = (".md", ".txt")) -> Path | None:
    """Newest snapshot dated on or before ``on_or_before``.

    Dated filenames let a backfill run see only what was knowable that day,
    which is the difference between a backtest and a look-ahead fantasy.
    """
    if not directory.exists():
        return None
    best: tuple[_dt.date, Path] | None = None
    for p in directory.iterdir():
        if p.suffix.lower() not in suffixes:
            continue
        try:
            d = _dt.date.fromisoformat(p.stem[:10])
        except ValueError:
            continue
        if d <= on_or_before and (best is None or d > best[0]):
            best = (d, p)
    return best[1] if best else None


def read_tracker(directory: Path) -> list[dict]:
    """Read the CSR handoff tracker from the newest .xlsx present."""
    if not directory.exists():
        return []
    files = sorted(directory.glob("*.xlsx"), key=lambda p: p.stat().st_mtime)
    if not files:
        return []
    try:
        import openpyxl
    except ImportError:
        print("[warn] openpyxl not installed; skipping tracker", file=sys.stderr)
        return []
    wb = openpyxl.load_workbook(files[-1], data_only=True)
    ws = next((wb[n] for n in wb.sheetnames if "track" in n.lower()), wb.worksheets[0])
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    out = []
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        out.append({header[i]: r[i] for i in range(min(len(header), len(r)))})
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--date", help="run as of this date (YYYY-MM-DD), default today")
    ap.add_argument("--root", default=str(ROOT))
    ap.add_argument("--dry-run", action="store_true", help="compute but write nothing")
    ap.add_argument("--quiet", action="store_true", help="suppress the brief on stdout")
    ap.add_argument("--json", action="store_true", help="emit the run JSON on stdout")
    ap.add_argument("--no-fetch", action="store_true",
                    help="skip the live snapshot endpoint; use on-disk only")
    ap.add_argument("--allow-stale", action="store_true",
                    help="build a brief even when the intake is not current. "
                         "Only for backfills and debugging: the figures will "
                         "describe the snapshot's date, not today's.")
    ap.add_argument("--max-age-hours", type=float, default=26.0,
                    help="refuse an intake older than this (default 26, which "
                         "is one Fiverr publishing cycle plus a margin)")
    args = ap.parse_args()

    root = Path(args.root)
    today = _dt.date.fromisoformat(args.date) if args.date else _dt.date.today()

    raw = root / "data" / "raw"
    snap_dir = raw / "snapshots"

    # ---- belt and braces -------------------------------------------------
    # Two independent producers: the Apps Script timer writing into
    # data/raw/snapshots/, and a live fetch of the web-app endpoint. Either
    # can be missing on any morning; take whichever is newer so one path
    # failing costs freshness rather than the whole run.
    live = disk = None
    if not args.no_fetch:
        try:
            live = snapshot.fetch_from_env()
            if live:
                print(f"[snapshot] fetched live, generated {live.generated_at}",
                      file=sys.stderr)
        except snapshot.SnapshotError as exc:
            print(f"[warn] live snapshot fetch failed: {exc}", file=sys.stderr)
    try:
        disk = snapshot.latest_on_disk(snap_dir, on_or_before=today)
    except snapshot.SnapshotError as exc:
        print(f"[warn] on-disk snapshot unreadable: {exc}", file=sys.stderr)

    snap = snapshot.pick_freshest(live, disk)
    src = None
    if snap:
        src = "live endpoint" if snap is live else "disk"
        print(f"[snapshot] using {src}, {snap.age_hours():.1f}h old, "
              f"{len(snap.tables)} tables", file=sys.stderr)
        for w in snap.warnings:
            print(f"[snapshot warn] {w}", file=sys.stderr)
        # Persist a freshly fetched snapshot so the next run has a fallback.
        if snap is live and not args.dry_run:
            snapshot.dump(snap, snap_dir / f"{today.isoformat()}.json")

    # ---- legacy markdown fallback ----------------------------------------
    orders_p = inq_p = None
    if snap is None:
        orders_p = latest_snapshot(raw / "orders", today)
        inq_p = latest_snapshot(raw / "inquiries", today)
        if not orders_p and not inq_p:
            print(f"[error] no snapshot and no markdown exports under {raw}. "
                  f"Deploy automation/Snapshot.gs or fetch the sheets — see "
                  f"docs/RUNBOOK.md.", file=sys.stderr)
            return 2
        print("[snapshot] none available; falling back to markdown exports",
              file=sys.stderr)

        # THE FALLBACK IS NOT AN EQUAL SUBSTITUTE, AND IT MUST SAY SO.
        #
        # Two things are true of a Drive markdown export of these workbooks
        # that are not true of a snapshot, and neither shows up in the output:
        #
        #   It gets truncated. `assert_complete` refuses that outright, below.
        #
        #   It carries no tab names. Both workbooks hold one tab per profile
        #   with identical headers, so eleven blocks arrive indistinguishable
        #   and this hub's profile cannot be separated from the other ten. On
        #   2026-08-06 that read 992 leads where the snapshot read 329, which
        #   would have put every conversion rate on the brief at a third of its
        #   real value with nothing on the page looking wrong.
        #
        # It is left usable because backfill depends on it and older exports
        # predate the multi-profile tabs, but it is never silent again.
        print("[snapshot warn] the markdown path cannot separate profiles: "
              "both workbooks keep one identically-headed tab per profile and "
              "the export drops the names. Treat every count below as covering "
              "more than this profile until a snapshot is available.",
              file=sys.stderr)

        for label, path in (("orders", orders_p), ("inquiries", inq_p)):
            if not path:
                continue
            try:
                ingest.assert_complete(path.read_text(encoding="utf-8"), label=str(path))
            except ingest.TruncatedExport as exc:
                print(f"[error] {exc}", file=sys.stderr)
                print(f"[error] refusing to build a brief on a partial {label} "
                      f"workbook. Re-export it, or restore a snapshot under "
                      f"{snap_dir}.", file=sys.stderr)
                return 2

    # ---- REFUSE TO BE WRONG, RATHER THAN TRY TO BE RIGHT --------------------
    #
    # The Routine fires every morning and this is the only thing standing
    # between "it worked" and "it produced a confident brief about last
    # Tuesday". Every other guard in this file catches a malformed input; this
    # one catches a well-formed input that is simply old.
    #
    # It matters because the failure is invisible by construction. A run on a
    # 44-hour-old snapshot produces a brief identical in every respect to a
    # live one, and on 2026-08-07 the difference between those two was the
    # whole verdict: BREACH at index 26 against HEALTHY at 79, because twelve
    # orders had landed in between.
    #
    # So: current data, or no brief and a non-zero exit. A scheduled job that
    # fails loudly gets fixed. One that succeeds quietly with old numbers does
    # not, because nothing ever says it needs fixing.
    if snap is not None and not args.allow_stale:
        age = snap.age_hours()
        if age > args.max_age_hours:
            print(f"[error] the freshest snapshot is {age:.1f}h old, over the "
                  f"{args.max_age_hours:.0f}h limit.", file=sys.stderr)
            if not os.environ.get("XSTUDIOZ_SNAPSHOT_URL"):
                print("[error] and XSTUDIOZ_SNAPSHOT_URL is not set in this "
                      "shell, so nothing tried to fetch anything fresher. "
                      "Export it and XSTUDIOZ_SNAPSHOT_TOKEN, then re-run.",
                      file=sys.stderr)
            else:
                print("[error] the endpoint is configured but did not answer "
                      "with anything newer. Check the Apps Script deployment.",
                      file=sys.stderr)
            print("[error] refusing to build a brief on stale data. Pass "
                  "--allow-stale if you specifically want one, and read every "
                  "figure in it as describing the snapshot's date.",
                  file=sys.stderr)
            return 2

    gig_p = latest_snapshot(raw / "gig", today, suffixes=(".json",))
    gig = json.loads(gig_p.read_text()) if gig_p else None

    # Which path won, recorded rather than merely printed. `src` is set above
    # from whichever of the live endpoint and the on-disk file was fresher.
    intake = {
        "path": "live" if src == "live endpoint" else "disk" if src == "disk" else "markdown",
        "endpoint_configured": bool(os.environ.get("XSTUDIOZ_SNAPSHOT_URL")),
        "orders_export": orders_p.name if orders_p else None,
        "inquiries_export": inq_p.name if inq_p else None,
        "gig_capture": gig_p.name if gig_p else None,
    }

    art = pipeline.run_daily(
        root=root, today=today,
        snap=snap, intake=intake,
        orders_text=orders_p.read_text(encoding="utf-8") if orders_p else "",
        inquiries_text=inq_p.read_text(encoding="utf-8") if inq_p else "",
        tracker_rows=read_tracker(raw / "order_tracker"),
        gig=gig,
        write=not args.dry_run)

    if args.json:
        print(json.dumps(art.as_dict(), indent=2, default=str))
    elif not args.quiet:
        print(art.brief_markdown)

    # Retired sheets get refused silently inside the ingester, which is the
    # right behaviour for the numbers and the wrong behaviour for the operator.
    # Say it on stderr every run it happens, next to the snapshot lines.
    ing = art.ingest_stats
    if ing.get("retired_tables_total"):
        print(f"[retired] refused {ing['retired_tables_total']} table(s), "
              f"{ing['retired_rows_total']} rows from sheets the hub replaced: "
              + ", ".join(f"{k}={v}" for k, v in
                          sorted(ing["retired_tables_skipped"].items())),
              file=sys.stderr)
        for r in ing.get("retired_tables") or []:
            print(f"[retired] {r['source_id']}/{r['name']}: {r['rows']} rows "
                  f"not counted: {r['why']}", file=sys.stderr)

    # Same voice for duplicate boards. Expected every run — the management
    # board lives in the workbook by design — so this is a receipt, not an
    # alarm: the operator can see the refusal happened and how many rows it
    # kept out of the book.
    for d in ing.get("duplicate_tables") or []:
        print(f"[duplicate] {d['source_id']}/{d['name']}: {d['rows']} rows "
              f"refused: {d['why']}", file=sys.stderr)

    print(f"\n[selfcheck] score={art.check.score:.0f}/100 "
          f"blocking={len(art.check.blocking_failures)} "
          f"emitted={art.emitted}", file=sys.stderr)
    for r in art.check.blocking_failures:
        print(f"[BLOCK] {r.name}: {r.detail}", file=sys.stderr)

    return 0 if art.emitted else 1


if __name__ == "__main__":
    raise SystemExit(main())
